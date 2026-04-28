#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
库存批次匹配工具
- SQLite 数据库存储（自动创建在 exe 同目录下）
- 导入 Excel 数据（支持 .xlsx 和 .xls）
- 按货物型号+库存匹配批次号组合（最优方案）
- 支持整列复制批次号/库存到 Excel
"""

import os
import sys
import sqlite3
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ==================== 路径 ====================
def get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

APP_DIR = get_app_dir()
DEFAULT_DB_PATH = os.path.join(APP_DIR, 'data.db')
DB_CONFIG_PATH = os.path.join(APP_DIR, 'db_config.json')
CURRENT_DB_PATH = DEFAULT_DB_PATH

def normalize_db_path(path):
    return os.path.abspath(os.path.expanduser(path))

def load_db_path_from_config():
    try:
        if not os.path.exists(DB_CONFIG_PATH):
            return DEFAULT_DB_PATH
        with open(DB_CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        saved_path = cfg.get('db_path', DEFAULT_DB_PATH)
        if not isinstance(saved_path, str) or not saved_path.strip():
            return DEFAULT_DB_PATH
        return normalize_db_path(saved_path)
    except Exception:
        return DEFAULT_DB_PATH

def save_db_path_to_config(path):
    cfg = {'db_path': normalize_db_path(path)}
    with open(DB_CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

def get_db_path():
    return CURRENT_DB_PATH

def set_db_path(path, persist=True):
    global CURRENT_DB_PATH
    CURRENT_DB_PATH = normalize_db_path(path)
    if persist:
        save_db_path_to_config(CURRENT_DB_PATH)

def connect_db():
    return sqlite3.connect(get_db_path(), timeout=30)

# ==================== 数据库 ====================
def init_db(db_path=None):
    if db_path:
        target = normalize_db_path(db_path)
    else:
        target = get_db_path()
    db_dir = os.path.dirname(target)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(target, timeout=30)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_no TEXT NOT NULL,
        category TEXT NOT NULL,
        qty INTEGER NOT NULL
    )''')
    c.execute('CREATE INDEX IF NOT EXISTS idx_category ON orders(category)')
    conn.commit()
    conn.close()

def db_add_records(records):
    conn = connect_db()
    c = conn.cursor()
    c.executemany('INSERT INTO orders (order_no, category, qty) VALUES (?, ?, ?)', records)
    conn.commit()
    conn.close()

def db_get_all():
    conn = connect_db()
    c = conn.cursor()
    c.execute('SELECT id, order_no, category, qty FROM orders ORDER BY id')
    rows = c.fetchall()
    conn.close()
    return rows

def db_get_by_category(category):
    conn = connect_db()
    c = conn.cursor()
    c.execute('SELECT id, order_no, category, qty FROM orders WHERE category = ? ORDER BY id', (category,))
    rows = c.fetchall()
    conn.close()
    return rows

def db_get_categories():
    conn = connect_db()
    c = conn.cursor()
    c.execute('SELECT DISTINCT category FROM orders ORDER BY category')
    cats = [r[0] for r in c.fetchall()]
    conn.close()
    return cats

def db_get_stats():
    conn = connect_db()
    c = conn.cursor()
    c.execute('SELECT COUNT(*), COALESCE(SUM(qty),0), COUNT(DISTINCT category) FROM orders')
    row = c.fetchone()
    conn.close()
    return row

def db_delete_ids(ids):
    if not ids:
        return
    conn = connect_db()
    c = conn.cursor()
    placeholders = ','.join('?' * len(ids))
    c.execute(f'DELETE FROM orders WHERE id IN ({placeholders})', ids)
    conn.commit()
    conn.close()

def db_consume_records(consumptions):
    """
    consumptions: [(id, consumed_qty), ...]
    若 consumed_qty >= 当前 qty，则删除该行；否则把 qty 减去 consumed_qty。
    用于支持"切件"消耗：同一批次记录可只领用部分库存。
    """
    if not consumptions:
        return
    conn = connect_db()
    c = conn.cursor()
    for rec_id, consumed in consumptions:
        if consumed <= 0:
            continue
        c.execute('SELECT qty FROM orders WHERE id = ?', (rec_id,))
        row = c.fetchone()
        if row is None:
            continue
        current_qty = row[0]
        new_qty = current_qty - consumed
        if new_qty <= 0:
            c.execute('DELETE FROM orders WHERE id = ?', (rec_id,))
        else:
            c.execute('UPDATE orders SET qty = ? WHERE id = ?', (new_qty, rec_id))
    conn.commit()
    conn.close()

def db_clear():
    conn = connect_db()
    c = conn.cursor()
    c.execute('DELETE FROM orders')
    try:
        c.execute("DELETE FROM sqlite_sequence WHERE name='orders'")
    except Exception:
        pass
    conn.commit()
    conn.close()

# ==================== Excel 解析（支持 .xlsx 和 .xls） ====================
def parse_excel(filepath):
    ext = os.path.splitext(filepath)[1].lower()

    if ext == '.xls':
        return _parse_xls(filepath)
    else:
        return _parse_xlsx(filepath)

def _parse_xlsx(filepath):
    try:
        import openpyxl
    except ImportError:
        raise ImportError("缺少 openpyxl 库，无法读取 .xlsx 文件")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(str(cell.value or '').strip())

    col_order = find_col(headers, ['批次号', '批次', '分单号', '单号', '订单号', '编号'])
    col_cat = find_col(headers, ['货物型号', '型号', '品类', '类别', '分类', '类型', '品名'])
    col_qty = find_col(headers, ['库存', '件数', '数量', '数目', '件'])

    if col_order is None or col_cat is None or col_qty is None:
        wb.close()
        raise ValueError(f"未找到必要列。当前表头：{headers}\n需要包含：批次号、货物型号、库存")

    records = []
    skipped = 0
    for row in ws.iter_rows(min_row=2):
        try:
            order_no = str(row[col_order].value or '').strip()
            category = str(row[col_cat].value or '').strip()
            qty_val = row[col_qty].value
            if qty_val is None:
                skipped += 1
                continue
            qty = int(float(str(qty_val)))
            if not order_no or not category or qty <= 0:
                skipped += 1
                continue
            records.append((order_no, category, qty))
        except (ValueError, TypeError, IndexError):
            skipped += 1

    wb.close()
    return records, skipped

def _parse_xls(filepath):
    try:
        import xlrd
    except ImportError:
        raise ImportError("缺少 xlrd 库，无法读取 .xls 文件。\n请运行：pip install xlrd")

    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        return [], 0

    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]

    col_order = find_col(headers, ['批次号', '批次', '分单号', '单号', '订单号', '编号'])
    col_cat = find_col(headers, ['货物型号', '型号', '品类', '类别', '分类', '类型', '品名'])
    col_qty = find_col(headers, ['库存', '件数', '数量', '数目', '件'])

    if col_order is None or col_cat is None or col_qty is None:
        raise ValueError(f"未找到必要列。当前表头：{headers}\n需要包含：批次号、货物型号、库存")

    records = []
    skipped = 0
    for r in range(1, ws.nrows):
        try:
            order_no = str(ws.cell_value(r, col_order)).strip()
            category = str(ws.cell_value(r, col_cat)).strip()
            qty_val = ws.cell_value(r, col_qty)
            if qty_val is None or qty_val == '':
                skipped += 1
                continue
            qty = int(float(str(qty_val)))
            if not order_no or not category or qty <= 0:
                skipped += 1
                continue
            records.append((order_no, category, qty))
        except (ValueError, TypeError, IndexError):
            skipped += 1

    return records, skipped

def find_col(headers, keywords):
    for i, h in enumerate(headers):
        for kw in keywords:
            if kw in h:
                return i
    return None

# ==================== 匹配算法 ====================
def _find_exact_subset(lst, target):
    """
    在 lst 中找一个**整批次精确凑出 target** 的子集，使用条数最少。
    每条记录最多使用一次（0/1 背包）。
    返回子集（lst 的子列表）或 None。
    """
    INF = float('inf')
    dp_count = [INF] * (target + 1)
    # dp_path[s] 保存到达和 s 时所选 item 的完整索引列表（不可变快照），
    # 避免后续迭代修改 dp_idx[s-q] 时污染已写入的 dp_path[s]，
    # 从而消除"同一条记录重复出现在结果中"的 bug。
    dp_path = [None] * (target + 1)
    dp_count[0] = 0
    dp_path[0] = []

    for i in range(len(lst)):
        q = lst[i][3]
        if q <= 0 or q > target:
            continue
        for s in range(target, q - 1, -1):
            prev_count = dp_count[s - q]
            if prev_count < INF and prev_count + 1 < dp_count[s]:
                dp_count[s] = prev_count + 1
                dp_path[s] = dp_path[s - q] + [i]

    if dp_count[target] == INF:
        return None
    return [lst[i] for i in reversed(dp_path[target])]


def find_best_match(items, target):
    """
    items: [(id, order_no, category, qty), ...]
    在某型号的全部批次中，找出使用条数最少的方案凑出 target 库存。
    每条批次记录最多使用一次；最后一条允许"切件"（仅领用部分库存）。

    返回 [(id, order_no, category, batch_qty, consumed_qty), ...] 或 None。
    consumed_qty <= batch_qty；当 consumed_qty < batch_qty 时表示切件。

    策略：
    1) 先用 partial-greedy（按库存量降序贪心）算出"允许切件"下的最少条数 k_partial。
    2) 再用 0/1 背包查"是否存在 k_partial 条整批次精确凑出 target"。
       - 若存在：返回整批次方案（不切件，数据更干净）。
       - 否则：返回 partial-greedy 方案（最后一条切件）。
    数学上 k_partial <= k_exact 恒成立，所以这两种方案的批次数相同。
    """
    n = min(len(items), 300)
    lst = items[:n]

    if target <= 0:
        return None

    total = sum(r[3] for r in lst if r[3] > 0)
    if total < target:
        return None

    # 1) partial-greedy：按 qty 降序取整批次直到累计 >= target，最后一条切件。
    sorted_lst = sorted([r for r in lst if r[3] > 0], key=lambda r: -r[3])
    cum = 0
    picked = []
    for r in sorted_lst:
        picked.append(r)
        cum += r[3]
        if cum >= target:
            break
    if cum < target:
        return None
    k_partial = len(picked)

    # 2) 尝试用同样的批次数找到整批次精确组合
    exact = _find_exact_subset(lst, target)
    if exact is not None and len(exact) == k_partial:
        return [(r[0], r[1], r[2], r[3], r[3]) for r in exact]

    # 3) 退化为切件方案：picked 的前 k_partial-1 条整批，最后一条切件
    overflow = cum - target
    last = picked[-1]
    last_consumed = last[3] - overflow
    result = [(r[0], r[1], r[2], r[3], r[3]) for r in picked[:-1]]
    result.append((last[0], last[1], last[2], last[3], last_consumed))
    return result

# ==================== GUI ====================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("库存批次匹配工具")
        self.geometry("960x750")
        self.minsize(800, 600)

        self.BG = "#f0f2f5"
        self.CARD = "#ffffff"
        self.PRIMARY = "#1a73e8"
        self.SUCCESS = "#0d9f6e"
        self.DANGER = "#e53e3e"

        self.configure(bg=self.BG)

        self._match_result = None
        # _match_consumptions: [(id, consumed_qty), ...]，用于删除时按实际消耗量更新数据库
        self._match_consumptions = []

        self._setup_styles()
        self._build_ui()
        self.update_db_label()
        self.refresh_all()

    def _setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('Treeview', font=('Microsoft YaHei', 10), rowheight=26)
        style.configure('Treeview.Heading', font=('Microsoft YaHei', 10, 'bold'))

    def _build_ui(self):
        # ====== Header ======
        header = tk.Frame(self, bg=self.PRIMARY, height=48)
        header.pack(fill='x')
        header.pack_propagate(False)
        tk.Label(header, text="库存批次匹配工具", bg=self.PRIMARY, fg='white',
                 font=('Microsoft YaHei', 15, 'bold')).pack(pady=9)

        # ====== PanedWindow ======
        paned = ttk.PanedWindow(self, orient='vertical')
        paned.pack(fill='both', expand=True, padx=12, pady=8)

        # ------ 上半部分 ------
        top_frame = tk.Frame(paned, bg=self.BG)
        paned.add(top_frame, weight=1)

        # 数据管理栏
        mgmt = tk.Frame(top_frame, bg=self.CARD, padx=12, pady=8)
        mgmt.pack(fill='x', pady=(0, 6))

        tk.Button(mgmt, text="导入 Excel", command=self.import_excel,
                  bg=self.PRIMARY, fg='white', font=('Microsoft YaHei', 10, 'bold'),
                  relief='flat', padx=14, pady=4, cursor='hand2').pack(side='left', padx=(0, 8))
        tk.Button(mgmt, text="刷新", command=self.refresh_all,
                  bg='#6c757d', fg='white', font=('Microsoft YaHei', 10),
                  relief='flat', padx=10, pady=4, cursor='hand2').pack(side='left', padx=(0, 8))
        tk.Button(mgmt, text="选择数据库", command=self.choose_database,
                  bg='#17a2b8', fg='white', font=('Microsoft YaHei', 10),
                  relief='flat', padx=10, pady=4, cursor='hand2').pack(side='left', padx=(0, 8))

        self.lbl_total = tk.Label(mgmt, text="记录：0", bg='#e8f0fe', fg=self.PRIMARY,
                                   font=('Microsoft YaHei', 9, 'bold'), padx=8, pady=2)
        self.lbl_total.pack(side='left', padx=(12, 6))
        self.lbl_cats = tk.Label(mgmt, text="型号：0", bg='#e8f0fe', fg=self.PRIMARY,
                                  font=('Microsoft YaHei', 9, 'bold'), padx=8, pady=2)
        self.lbl_cats.pack(side='left', padx=(0, 6))
        self.lbl_qty = tk.Label(mgmt, text="总库存：0", bg='#e8f0fe', fg=self.PRIMARY,
                                 font=('Microsoft YaHei', 9, 'bold'), padx=8, pady=2)
        self.lbl_qty.pack(side='left')

        tk.Button(mgmt, text="清空所有", command=self.clear_all,
                  bg=self.DANGER, fg='white', font=('Microsoft YaHei', 9),
                  relief='flat', padx=10, pady=4, cursor='hand2').pack(side='right')

        # 筛选
        filter_bar = tk.Frame(top_frame, bg=self.CARD, padx=12, pady=4)
        filter_bar.pack(fill='x', pady=(0, 2))
        tk.Label(filter_bar, text="货物型号筛选：", bg=self.CARD, font=('Microsoft YaHei', 10)).pack(side='left')
        self.filter_var = tk.StringVar(value='全部')
        self.filter_combo = ttk.Combobox(filter_bar, textvariable=self.filter_var,
                                          state='readonly', width=18, font=('Microsoft YaHei', 10))
        self.filter_combo.pack(side='left', padx=6)
        self.filter_combo.bind('<<ComboboxSelected>>', lambda e: self.refresh_table())

        self.lbl_db = tk.Label(filter_bar, text="", bg=self.CARD,
                                fg='#aaa', font=('Microsoft YaHei', 8))
        self.lbl_db.pack(side='right')

        # 数据表
        tree_frame = tk.Frame(top_frame, bg=self.CARD)
        tree_frame.pack(fill='both', expand=True)

        cols = ('no', 'order_no', 'category', 'qty')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings', height=8)
        self.tree.heading('no', text='序号')
        self.tree.heading('order_no', text='批次号')
        self.tree.heading('category', text='货物型号')
        self.tree.heading('qty', text='库存')
        self.tree.column('no', width=55, anchor='center')
        self.tree.column('order_no', width=280)
        self.tree.column('category', width=140, anchor='center')
        self.tree.column('qty', width=80, anchor='center')

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        vsb.pack(side='right', fill='y')

        # ------ 下半部分：匹配查询 + 结果 ------
        bot_frame = tk.Frame(paned, bg=self.BG)
        paned.add(bot_frame, weight=1)

        # 查询栏
        query_bar = tk.Frame(bot_frame, bg=self.CARD, padx=12, pady=8)
        query_bar.pack(fill='x', pady=(0, 6))

        tk.Label(query_bar, text="货物型号：", bg=self.CARD, font=('Microsoft YaHei', 10)).pack(side='left')
        self.query_cat_var = tk.StringVar()
        self.query_cat_combo = ttk.Combobox(query_bar, textvariable=self.query_cat_var,
                                             state='readonly', width=14, font=('Microsoft YaHei', 10))
        self.query_cat_combo.pack(side='left', padx=(0, 14))

        tk.Label(query_bar, text="目标库存：", bg=self.CARD, font=('Microsoft YaHei', 10)).pack(side='left')
        self.query_qty_var = tk.StringVar()
        tk.Entry(query_bar, textvariable=self.query_qty_var, width=10,
                 font=('Microsoft YaHei', 10)).pack(side='left', padx=(0, 14))

        tk.Button(query_bar, text="开始匹配", command=self.do_match,
                  bg=self.PRIMARY, fg='white', font=('Microsoft YaHei', 10, 'bold'),
                  relief='flat', padx=16, pady=4, cursor='hand2').pack(side='left')

        # 结果区域
        result_container = tk.Frame(bot_frame, bg=self.CARD, padx=12, pady=8)
        result_container.pack(fill='both', expand=True)

        self.result_title = tk.Label(result_container, text="", bg=self.CARD,
                                      fg=self.PRIMARY, font=('Microsoft YaHei', 11, 'bold'))
        self.result_title.pack(anchor='w')

        # 结果表格
        res_tree_frame = tk.Frame(result_container, bg=self.CARD)
        res_tree_frame.pack(fill='both', expand=True, pady=(4, 6))

        res_cols = ('r_no', 'r_order', 'r_qty')
        self.res_tree = ttk.Treeview(res_tree_frame, columns=res_cols, show='headings', height=6)
        self.res_tree.heading('r_no', text='序号')
        self.res_tree.heading('r_order', text='批次号')
        self.res_tree.heading('r_qty', text='库存')
        self.res_tree.column('r_no', width=55, anchor='center')
        self.res_tree.column('r_order', width=360)
        self.res_tree.column('r_qty', width=100, anchor='center')

        res_vsb = ttk.Scrollbar(res_tree_frame, orient='vertical', command=self.res_tree.yview)
        self.res_tree.configure(yscrollcommand=res_vsb.set)
        self.res_tree.pack(side='left', fill='both', expand=True)
        res_vsb.pack(side='right', fill='y')

        # 操作按钮栏
        btn_bar = tk.Frame(result_container, bg=self.CARD)
        btn_bar.pack(fill='x', pady=(0, 2))

        tk.Button(btn_bar, text="复制批次号列", command=self.copy_order_col,
                  bg='#ffffff', fg=self.PRIMARY, font=('Microsoft YaHei', 10, 'bold'),
                  relief='solid', bd=1, padx=14, pady=4, cursor='hand2').pack(side='left', padx=(0, 8))

        tk.Button(btn_bar, text="复制库存列", command=self.copy_qty_col,
                  bg='#ffffff', fg=self.PRIMARY, font=('Microsoft YaHei', 10, 'bold'),
                  relief='solid', bd=1, padx=14, pady=4, cursor='hand2').pack(side='left', padx=(0, 16))

        tk.Button(btn_bar, text="仅删除记录", command=self.delete_only,
                  bg=self.DANGER, fg='white', font=('Microsoft YaHei', 10),
                  relief='flat', padx=14, pady=4, cursor='hand2').pack(side='left')

        # 状态提示栏（替代弹窗）
        self.status_label = tk.Label(result_container, text="", bg=self.CARD,
                                      fg=self.SUCCESS, font=('Microsoft YaHei', 10, 'bold'))
        self.status_label.pack(anchor='w', pady=(4, 0))
        self._status_timer = None

    # ==================== 数据操作 ====================
    def import_excel(self):
        filepath = filedialog.askopenfilename(
            title="选择要导入的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not filepath:
            return
        try:
            records, skipped = parse_excel(filepath)
            if not records:
                messagebox.showwarning("提示", "文件中没有有效数据可导入")
                return
            db_add_records(records)
            msg = f"成功导入 {len(records)} 条记录"
            if skipped:
                msg += f"\n跳过 {skipped} 条无效数据"
            messagebox.showinfo("导入成功", msg)
            self.refresh_all()
        except ImportError as e:
            messagebox.showerror("错误", str(e))
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def _short_db_path(self, path, max_len=72):
        if len(path) <= max_len:
            return path
        head = (max_len - 3) // 2
        tail = max_len - 3 - head
        return f"{path[:head]}...{path[-tail:]}"

    def update_db_label(self):
        self.lbl_db.config(text=f"数据库：{self._short_db_path(get_db_path())}")

    def choose_database(self):
        current_path = get_db_path()
        initial_dir = os.path.dirname(current_path) or APP_DIR
        db_path = filedialog.asksaveasfilename(
            title="选择或新建数据库文件",
            initialdir=initial_dir,
            initialfile=os.path.basename(current_path),
            defaultextension=".db",
            filetypes=[("SQLite 数据库", "*.db *.sqlite *.sqlite3"), ("所有文件", "*.*")]
        )
        if not db_path:
            return
        db_path = normalize_db_path(db_path)
        try:
            init_db(db_path)
            set_db_path(db_path, persist=True)
        except Exception as e:
            messagebox.showerror("切换失败", f"无法使用该数据库文件：\n{db_path}\n\n{e}")
            return

        self.clear_results()
        self.update_db_label()
        self.refresh_all()
        messagebox.showinfo("切换成功", f"当前数据库：\n{db_path}")

    def clear_all(self):
        if not messagebox.askyesno("确认", "确定清空所有数据吗？"):
            return
        db_clear()
        self.clear_results()
        self.refresh_all()

    def refresh_all(self):
        cats = db_get_categories()
        stats = db_get_stats()

        self.lbl_total.config(text=f"记录：{stats[0]}")
        self.lbl_cats.config(text=f"型号：{stats[2]}")
        self.lbl_qty.config(text=f"总库存：{stats[1]}")

        filter_vals = ['全部'] + cats
        self.filter_combo['values'] = filter_vals
        if self.filter_var.get() not in filter_vals:
            self.filter_var.set('全部')

        self.query_cat_combo['values'] = cats
        if self.query_cat_var.get() not in cats:
            self.query_cat_var.set('')

        self.refresh_table()

    def refresh_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        cat = self.filter_var.get()
        rows = db_get_all() if cat == '全部' else db_get_by_category(cat)

        for i, row in enumerate(rows):
            tag = 'even' if i % 2 == 0 else 'odd'
            self.tree.insert('', 'end', values=(i + 1, row[1], row[2], row[3]), tags=(tag,))

        self.tree.tag_configure('even', background='#ffffff')
        self.tree.tag_configure('odd', background='#f8f9fa')

    # ==================== 匹配 ====================
    def do_match(self):
        category = self.query_cat_var.get()
        qty_str = self.query_qty_var.get()

        if not category:
            messagebox.showwarning("提示", "请选择货物型号")
            return
        try:
            target_qty = int(qty_str)
            if target_qty <= 0:
                raise ValueError
        except (ValueError, TypeError):
            messagebox.showwarning("提示", "请输入有效的目标库存（正整数）")
            return

        items = db_get_by_category(category)
        if not items:
            messagebox.showinfo("结果", "该型号下没有数据")
            return

        total_avail = sum(r[3] for r in items)
        if total_avail < target_qty:
            messagebox.showinfo("结果", f"该型号下总库存为 {total_avail}，不足目标库存 {target_qty}")
            return

        match = find_best_match(items, target_qty)
        if not match:
            messagebox.showinfo("结果",
                f"无法凑出 {target_qty} 件库存\n"
                f"该型号共 {len(items)} 条记录，总库存 {total_avail}")
            return

        self.show_result(match)

    # ==================== 结果展示 ====================
    def clear_results(self):
        self._match_result = None
        self._match_consumptions = []
        self.result_title.config(text="")
        for item in self.res_tree.get_children():
            self.res_tree.delete(item)

    def show_result(self, match):
        # match: [(id, order_no, category, batch_qty, consumed_qty), ...]
        self._match_result = match
        self._match_consumptions = [(r[0], r[4]) for r in match]

        sum_consumed = sum(r[4] for r in match)
        partial_count = sum(1 for r in match if r[4] < r[3])
        if partial_count:
            title = (f"匹配结果：{len(match)} 个批次号（含 {partial_count} 个切件），"
                     f"总库存 {sum_consumed}")
        else:
            title = f"匹配结果：{len(match)} 个批次号，总库存 {sum_consumed}"
        self.result_title.config(text=title)

        for item in self.res_tree.get_children():
            self.res_tree.delete(item)

        for i, r in enumerate(match):
            batch_qty = r[3]
            consumed_qty = r[4]
            if consumed_qty < batch_qty:
                # 切件：同时呈现领用数与原批库存，便于用户核对
                qty_text = f"{consumed_qty} / {batch_qty}"
            else:
                qty_text = str(consumed_qty)
            tag = 'even' if i % 2 == 0 else 'odd'
            self.res_tree.insert('', 'end', values=(i + 1, r[1], qty_text), tags=(tag,))

        self.res_tree.tag_configure('even', background='#ffffff')
        self.res_tree.tag_configure('odd', background='#f0f4ff')

    # ==================== 状态提示（替代弹窗） ====================
    def show_status(self, msg, color=None, duration=3000):
        if color is None:
            color = self.SUCCESS
        self.status_label.config(text=msg, fg=color)
        if self._status_timer:
            self.after_cancel(self._status_timer)
        self._status_timer = self.after(duration, lambda: self.status_label.config(text=""))

    # ==================== 复制功能 ====================
    def _check_result(self):
        if not self._match_result:
            self.show_status("请先执行匹配查询", color=self.DANGER)
            return False
        return True

    def copy_order_col(self):
        if not self._check_result():
            return
        text = '\n'.join(r[1] for r in self._match_result)
        self.clipboard_clear()
        self.clipboard_append(text)
        self.show_status(f"已复制 {len(self._match_result)} 个批次号")

    def copy_qty_col(self):
        if not self._check_result():
            return
        # 库存列复制的是本次领用量（切件后的实际领用数），而非原批库存
        text = '\n'.join(str(r[4]) for r in self._match_result)
        self.clipboard_clear()
        self.clipboard_append(text)
        n = len(self._match_result)
        self.show_status(f"已复制 {n} 个库存")

        partial_count = sum(1 for r in self._match_result if r[4] < r[3])
        if partial_count:
            extra = (f"\n其中 {partial_count} 个为切件领用，"
                     f"删除时全领用的批次会被删行，"
                     f"切件的批次只会从原库存中减去领用量。")
        else:
            extra = ""

        # 复制完成后立即提醒用户是否删除已匹配的记录，避免漏删导致库存被重复领用
        if messagebox.askyesno(
            "是否删除记录",
            f"已复制 {n} 个库存。\n是否同时从数据库中扣减本次领用量？{extra}\n\n"
            "是 → 扣减 / 删除，避免下次再被重复领用\n"
            "否 → 保留，可稍后再点\"仅删除记录\""
        ):
            consumptions = list(self._match_consumptions)
            db_consume_records(consumptions)
            self.clear_results()
            self.refresh_all()
            self.show_status(f"已复制并领用 {len(consumptions)} 条记录")

    def delete_only(self):
        if not self._check_result():
            return
        n = len(self._match_consumptions)
        db_consume_records(list(self._match_consumptions))
        self.clear_results()
        self.refresh_all()
        self.show_status(f"已领用 {n} 条记录")

# ==================== Main ====================
if __name__ == '__main__':
    set_db_path(load_db_path_from_config(), persist=False)
    init_db()
    app = App()
    app.mainloop()
